import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ==========================================
# 1. KONFIGURASI HALAMAN & SISTEM KEAMANAN
# ==========================================
st.set_page_config(
    page_title="SPX Data Processor",
    page_icon="📦",
    layout="wide"
)

def check_password():
    """Mengembalikan True jika kata sandi benar."""
    def password_entered():
        target_password = st.secrets.get("APP_PASSWORD", "deon22")
        if st.session_state["password"] == target_password:
            st.session_state["password_correct"] = True
            del st.session_state["password"]
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state or not st.session_state["password_correct"]:
        st.title("🔒 Area Terbatas")
        st.text_input(
            "Masukkan Kata Sandi Aplikasi:", 
            type="password", 
            on_change=password_entered, 
            key="password"
        )
        if st.session_state.get("password_correct") == False:
            st.error("😕 Kata sandi salah. Silakan coba lagi.")
        return False
    return True

if not check_password():
    st.stop()

# ==========================================
# 2. MASTER MAP & STYLING EXCEL
# ==========================================
MARKING_MAP = {
    'Abepura DC': 'DJJ', 'Alak DC': 'KOE', 'Bacan Hub': 'LAH', 'Baguala DC': 'AMQ',
    'Balikpapan DC': 'BPN', 'Banjarmasin DC': 'BDJ1', 'Banjarmasin 2 DC': 'BDJ2',
    'Banjarbaru DC': 'BJB', 'Batam DC': 'BTH', 'Dungingi DC': 'DGI',
    'Kalawat DC': 'MDU', 'Kota Waingapu Hub': 'WGP', 'Kota Waingapu 2 Hub': 'WGP2',
    'Kota Waingapu 4 Hub': 'WGP4', 'Labuhan Bajo DC': 'LBJ', 'Loli Hub': 'TMC2',
    'Loura (Laura) Hub': 'TMC', 'Manokwari Barat DC': 'MKW', 'Mantikulore DC': 'PLW',
    'Medan DC': 'KNO', 'Medan Amplas DC': 'KNO2', 'Merauke DC': 'MKQ',
    'Mimika Baru Hub': 'TIM', 'Nabire Hub': 'NBX', 'Palangka Raya DC': 'PKY',
    'Percut Sei Tuan DC': 'PST', 'Pekanbaru DC': 'PKU', 'Pekanbaru 2 DC': 'PKU2',
    'Pontianak DC': 'PNK', 'Pontianak 2 DC': 'PNK2', 'Sungai Kakap DC': 'PNK3',
    'Sorong Utara DC': 'SOQ', 'Tarakan Barat Hub': 'TRKB', 'Tarakan Barat 4 Hub': 'TRKB4',
    'Tarakan Timur Hub': 'TRKT', 'Tarakan Utara Hub': 'TRKU', 'Teluk Mutiara Hub': 'ARD',
    'Ternate Hub': 'TTE', 'Ternate Utara Hub': 'TTU', 'Ternate Selatan Hub': 'TTS',
    'Ternate Selatan 2 Hub': 'TTS2', 'Ternate Selatan 3 Hub': 'TTS3',
    'Wamena Hub': 'WMX', 'Wua-Wua DC': 'KDI', 'Medan Deli DC': 'KNO3'
}

def get_prefix_code(destination_name):
    if pd.isna(destination_name): return "UNK"
    dest_str = str(destination_name).strip()
    return MARKING_MAP.get(dest_str, dest_str.split()[0][:3].upper())

RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
GRAY_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BLUE_SJM_FILL = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

FONT_WHITE_TITLE_18 = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_SUB_BLACK = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="000000")
FONT_GRAND_TOTAL = Font(name="Calibri", size=11, bold=True, color="000000")
FONT_REGULAR_BLACK = Font(name="Calibri", size=9, color="000000")
FONT_BIG_TOTAL = Font(name="Calibri", size=20, bold=True, color="000000")

BORDER_THIN = Border(
    left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000')
)

ALIGN_FULL_CENTER = Alignment(horizontal='center', vertical='center')

def autofit_table_columns(ws, start_row=1, min_width=15):
    for col in ws.iter_cols(min_row=start_row, max_row=ws.max_row):
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in col:
            if type(cell).__name__ == 'MergedCell':
                continue
            val = cell.value
            if val is None:
                continue
            if isinstance(val, (float, int)):
                val_str = f"{val:,.2f}"
            else:
                val_str = str(val)
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 8, min_width)

# ==========================================
# 3. FUNGSI PEMROSESAN DATA
# ==========================================
def process_excel_data(uploaded_file):
    df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    wb = openpyxl.Workbook()

    # Cek kecukupan baris
    if len(df_raw) < 4:
        raise ValueError("File Excel tidak memiliki cukup baris data (minimal 4 baris).")

    df_data = df_raw.iloc[3:, :8].copy()
    df_data.columns = ['Tanggal', 'Vendor', 'Sc_Origin', 'Sc_Destination', 'Lt_Number', 'To_Number', 'Gross_Weight', 'Total']

    # Filter data valid
    df_data = df_data[df_data['To_Number'].notna()].copy()
    if df_data.empty:
        raise ValueError("Tidak ditemukan data transaksi yang memiliki 'To_Number' (Kolom F) di baris 4 ke bawah.")

    # Format Tanggal Transaksi
    df_data['Tanggal'] = pd.to_datetime(df_data['Tanggal'], errors='coerce').dt.strftime('%Y-%m-%d')

    # Format Gross Weight aman (ubah koma ke titik bila berupa string)
    df_data['Gross_Weight'] = df_data['Gross_Weight'].astype(str).str.replace(',', '.')
    df_data['Gross_Weight'] = pd.to_numeric(df_data['Gross_Weight'], errors='coerce').fillna(0.0)

    # Sorting
    df_reversed = df_data.iloc[::-1].copy()
    df_sorted = df_reversed.sort_values(by='Sc_Destination', kind='stable', ascending=True).reset_index(drop=True)
    
    df_sorted['to_index'] = df_sorted.groupby('Sc_Destination').cumcount()
    df_sorted['bag_num'] = (df_sorted['to_index'] // 15) + 1

    # 1. SHEET 'SJM'
    ws_sjm = wb.active
    ws_sjm.title = "SJM"

    title_text = str(df_raw.iloc[0, 0]) if (not pd.isna(df_raw.iloc[0, 0])) else "SURAT JALAN MANUAL"

    # Perbaikan Sub-Title Tanggal Aman (Safe Parsing)
    sub_title_raw = df_raw.iloc[1, 0] if len(df_raw) > 1 else ""
    sub_title_text = ""
    if pd.notna(sub_title_raw):
        parsed_date = pd.to_datetime(sub_title_raw, errors='coerce')
        if pd.notnull(parsed_date):
            sub_title_text = parsed_date.strftime('%d %B %Y').upper() + " TRIP 1"
        else:
            sub_title_text = str(sub_title_raw)

    code_box = str(df_raw.iloc[0, 7]) if (df_raw.shape[1] >= 8 and pd.notna(df_raw.iloc[0, 7])) else ""

    ws_sjm.append([title_text, "", "", "", "", "", "", code_box])
    ws_sjm.append([sub_title_text, "", "", "", "", "", "", ""])

    ws_sjm.merge_cells("A1:G1")
    ws_sjm.merge_cells("A2:G2")

    for r in [1, 2]:
        for c in range(1, 8):
            cell = ws_sjm.cell(row=r, column=c)
            cell.fill = BLUE_SJM_FILL
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_FULL_CENTER
            cell.border = BORDER_THIN

    ws_sjm.cell(1, 8).fill = BLUE_SJM_FILL
    ws_sjm.cell(1, 8).font = FONT_HEADER
    ws_sjm.cell(1, 8).alignment = ALIGN_FULL_CENTER
    ws_sjm.cell(1, 8).border = BORDER_THIN

    ws_sjm.cell(2, 8).value = len(df_sorted)
    ws_sjm.cell(2, 8).font = FONT_BIG_TOTAL
    ws_sjm.cell(2, 8).alignment = ALIGN_FULL_CENTER
    ws_sjm.cell(2, 8).border = BORDER_THIN

    headers_sjm = ['TGL', 'Vendor', 'SC Orgin', 'DESTINATION', 'LT NUMBER', 'TO NUMBER', 'Gross Weight', 'TOTAL']
    ws_sjm.append(headers_sjm)
    for c_idx in range(1, 9):
        cell = ws_sjm.cell(row=3, column=c_idx)
        cell.fill = GRAY_HEADER_FILL
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_FULL_CENTER
        cell.border = BORDER_THIN

    for row in df_sorted.itertuples():
        ws_sjm.append([row.Tanggal, row.Vendor, row.Sc_Origin, row.Sc_Destination, row.Lt_Number, row.To_Number, row.Gross_Weight, ""])

    for row in ws_sjm.iter_rows(min_row=4, max_row=ws_sjm.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = BORDER_THIN
            cell.font = FONT_REGULAR_BLACK
            cell.alignment = ALIGN_FULL_CENTER
            if isinstance(cell.value, (float, int)):
                cell.number_format = '#,##0.00'

    tot_sjm_row = ws_sjm.max_row + 1
    total_sjm_gw = round(df_sorted['Gross_Weight'].sum(), 2)
    ws_sjm.append(["TOTAL", "", "", "", "", "", total_sjm_gw, ""])
    ws_sjm.merge_cells(start_row=tot_sjm_row, start_column=1, end_row=tot_sjm_row, end_column=6)

    for c_idx in range(1, 9):
        cell = ws_sjm.cell(row=tot_sjm_row, column=c_idx)
        cell.fill = BLUE_SJM_FILL
        cell.font = FONT_HEADER
        cell.border = BORDER_THIN
        cell.alignment = ALIGN_FULL_CENTER
        if c_idx == 7:
            cell.number_format = '#,##0.00'

    autofit_table_columns(ws_sjm, start_row=3, min_width=16)

    # 2. SHEET 'MARKING'
    ws_marking = wb.create_sheet(title="MARKING")

    ws_marking.append(["MARKING SPX OSO SUB DC CYCLE 1 HB"] + [""] * 10)
    ws_marking.merge_cells("A1:K1")
    ws_marking.row_dimensions[1].height = 28

    cell_a1 = ws_marking["A1"]
    cell_a1.fill = RED_FILL
    cell_a1.font = FONT_WHITE_TITLE_18
    cell_a1.alignment = ALIGN_FULL_CENTER

    ws_marking.append([sub_title_text] + [""] * 10)
    ws_marking.merge_cells("A2:K2")
    ws_marking.row_dimensions[2].height = 20

    cell_a2 = ws_marking["A2"]
    cell_a2.fill = YELLOW_FILL
    cell_a2.font = FONT_SUB_BLACK
    cell_a2.alignment = ALIGN_FULL_CENTER

    headers_m = [
        "Tanggal", "Vendor", "Sc Origin", "Sc Destination", "Lt Number",
        "To Number", "Marking", "Gross Weight", "Remarks", "External Number", "Clear Gw"
    ]
    ws_marking.append(headers_m)
    ws_marking.row_dimensions[3].height = 22

    for col_idx in range(1, 12):
        c = ws_marking.cell(row=3, column=col_idx)
        c.fill = GRAY_HEADER_FILL
        c.font = FONT_HEADER
        c.alignment = ALIGN_FULL_CENTER
        c.border = BORDER_THIN

    marking_rows = []
    for idx, row in enumerate(df_sorted.itertuples(), start=4):
        tgl, vendor, origin, dest = row.Tanggal, "Lion Parcel", row.Sc_Origin, row.Sc_Destination
        lt_num, to_num, gw = row.Lt_Number, row.To_Number, row.Gross_Weight

        prefix = get_prefix_code(dest)
        marking = f"{prefix}-C1-{row.bag_num}"
        remarks = "BAG"

        formula_ext_num = f'=G{idx}&"/"&E{idx}&"/"&I{idx}'
        ext_num_val_for_df = f"{marking}/{lt_num}/{remarks}"

        marking_rows.append([tgl, vendor, origin, dest, lt_num, to_num, marking, gw, remarks, ext_num_val_for_df, gw, idx])
        ws_marking.append([tgl, vendor, origin, dest, lt_num, to_num, marking, gw, remarks, formula_ext_num, gw])

    for row in ws_marking.iter_rows(min_row=4, max_row=ws_marking.max_row, min_col=1, max_col=11):
        for c in row:
            c.border = BORDER_THIN
            c.font = FONT_REGULAR_BLACK
            c.alignment = ALIGN_FULL_CENTER
            if isinstance(c.value, (float, int)):
                c.number_format = '#,##0.00'

    df_m = pd.DataFrame(marking_rows, columns=headers_m + ['marking_row_idx'])
    df_m["Clear Gw"] = pd.to_numeric(df_m["Clear Gw"], errors='coerce').fillna(0)
    df_m["Gross Weight"] = pd.to_numeric(df_m["Gross Weight"], errors='coerce').fillna(0)

    max_marking_data_row = ws_marking.max_row

    tot_gw = round(df_m["Clear Gw"].sum(), 2)
    tot_row_idx = ws_marking.max_row + 1
    ws_marking.append(["GRAND TOTAL", "", "", "", "", "", "", tot_gw, "", "", tot_gw])

    ws_marking.merge_cells(start_row=tot_row_idx, start_column=1, end_row=tot_row_idx, end_column=7)
    for col_idx in range(1, 12):
        c = ws_marking.cell(row=tot_row_idx, column=col_idx)
        c.fill = RED_FILL
        c.font = FONT_GRAND_TOTAL
        c.border = BORDER_THIN
        c.alignment = ALIGN_FULL_CENTER
        if col_idx in [8, 11]:
            c.number_format = '#,##0.00'

    autofit_table_columns(ws_marking, start_row=3, min_width=18)

    # 3. SHEET 'PVT'
    ws_pvt = wb.create_sheet(title="PVT")
    ws_pvt.append([])
    ws_pvt.append([])
    ws_pvt.append(["External Number", "Count of External Number", "Sum of Clear Gw"])

    for c_idx in range(1, 4):
        c = ws_pvt.cell(row=3, column=c_idx)
        c.fill = GRAY_HEADER_FILL
        c.font = FONT_HEADER
        c.alignment = ALIGN_FULL_CENTER
        c.border = BORDER_THIN

    pvt_first_rows = df_m.drop_duplicates(subset=["External Number"], keep="first")

    for pvt_row_idx, item in enumerate(pvt_first_rows.itertuples(), start=4):
        m_row = item.marking_row_idx
        formula_pvt_ext = f"=MARKING!J{m_row}"
        formula_count = f"=COUNTIF(MARKING!J$4:J${max_marking_data_row}, A{pvt_row_idx})"
        formula_sum = f"=SUMIF(MARKING!J$4:J${max_marking_data_row}, A{pvt_row_idx}, MARKING!K$4:K${max_marking_data_row})"

        ws_pvt.append([formula_pvt_ext, formula_count, formula_sum])

    pvt_last_row = ws_pvt.max_row
    tot_pvt_count_formula = f"=SUM(B4:B{pvt_last_row})"
    tot_pvt_sum_formula = f"=SUM(C4:C{pvt_last_row})"

    ws_pvt.append(["Grand Total", tot_pvt_count_formula, tot_pvt_sum_formula])

    for row in ws_pvt.iter_rows(min_row=4, max_row=ws_pvt.max_row, min_col=1, max_col=3):
        is_last = (row[0].row == ws_pvt.max_row)
        for c in row:
            c.border = BORDER_THIN
            c.font = FONT_GRAND_TOTAL if is_last else FONT_REGULAR_BLACK
            c.alignment = ALIGN_FULL_CENTER
            if is_last: c.fill = GRAY_HEADER_FILL
            if c.column == 2:
                c.number_format = '#,##0'
            elif c.column == 3:
                c.number_format = '#,##0.00'

    for col in ws_pvt.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_pvt.column_dimensions[col_letter].width = max(max_len + 12, 30)

    # 4. SHEET 'Sheet3'
    ws_sheet3 = wb.create_sheet(title="Sheet3")
    ws_sheet3.append([])
    ws_sheet3.append([])
    ws_sheet3.append(["Sc Destination", "Count of To Number", "Sum of Gross Weight"])

    for c_idx in range(1, 4):
        c = ws_sheet3.cell(row=3, column=c_idx)
        c.fill = GRAY_HEADER_FILL
        c.font = FONT_HEADER
        c.alignment = ALIGN_FULL_CENTER
        c.border = BORDER_THIN

    dest_summary = df_m.groupby("Sc Destination", sort=False).agg(
        Count_TO=("To Number", "count"), Sum_Gw=("Gross Weight", "sum")
    ).reset_index()

    for row in dest_summary.itertuples(index=False):
        ws_sheet3.append([row[0], int(row[1]), round(row[2], 2)])

    ws_sheet3.append(["Grand Total", int(dest_summary["Count_TO"].sum()), round(dest_summary["Sum_Gw"].sum(), 2)])

    for row in ws_sheet3.iter_rows(min_row=4, max_row=ws_sheet3.max_row, min_col=1, max_col=3):
        is_last = (row[0].row == ws_sheet3.max_row)
        for c in row:
            c.border = BORDER_THIN
            c.font = FONT_GRAND_TOTAL if is_last else FONT_REGULAR_BLACK
            c.alignment = ALIGN_FULL_CENTER
            if is_last: c.fill = GRAY_HEADER_FILL
            if c.column == 2:
                c.number_format = '#,##0'
            elif c.column == 3:
                c.number_format = '#,##0.00'

    for col in ws_sheet3.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_sheet3.column_dimensions[col_letter].width = max(max_len + 12, 28)

    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    
    return output_stream
# ==========================================
# 4. ANTARMUKA UTAMA (MAIN APP UI)
# ==========================================
st.title("📦 Lion parcel Data Formatting & Marking Generator")
st.markdown("Unggah file Excel raw data Anda di bawah ini untuk menghasilkan file Excel dengan sheet **SJM**, **MARKING**, **PVT**, dan **Sheet3**.")

with st.sidebar:
    st.write("🔓 **Sesi Login Aktif**")
    if st.button("Keluar (Logout)"):
        st.session_state["password_correct"] = False
        st.rerun()

uploaded_file = st.file_uploader("Pilih file Excel (.xlsx / .xls)", type=["xlsx", "xls"])

if uploaded_file is not None:
    st.success(f"File **{uploaded_file.name}** berhasil diunggah!")
    
    if st.button("🚀 Proses Data Excel", type="primary"):
        with st.spinner("Sedang memproses data dan menyusun sheet Excel..."):
            try:
                processed_excel = process_excel_data(uploaded_file)
                output_filename = f"FIXED_SCRIPT_{uploaded_file.name}"
                
                
                st.success("Pemrosesan Selesai! Klik tombol di bawah untuk mengunduh hasil.")
                
                st.download_button(
                    label="📥 Unduh File Excel Hasil Processing",
                    data=processed_excel,
                    file_name=output_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Terjadi kesalahan saat memproses data: {e}")
